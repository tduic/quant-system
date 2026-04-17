"""FastAPI dashboard with 7 tabs + Excel export.

Each tab has its own GET endpoint returning JSON. The Excel export
bundles all tabs into a single workbook with separate sheets.

Symbol filtering: all endpoints accept an optional `?symbol=` query
parameter to filter data to a single symbol. Omit for portfolio-wide view.

Tab 7 (Backtest Analysis) uses async job submission — POST to submit,
GET to poll status, GET to retrieve results.
"""

from __future__ import annotations

import io
import logging
from typing import TYPE_CHECKING

from fastapi import FastAPI, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from post_trade_svc.analysis_jobs import JobStore, _list_historical_backtests

if TYPE_CHECKING:
    from post_trade_svc.state import PostTradeState


class AnalysisRequest(BaseModel):
    """Request body for submitting an analysis job."""

    analysis_type: str
    params: dict = {}


logger = logging.getLogger(__name__)


def create_app(state: PostTradeState) -> FastAPI:
    """Create the FastAPI app with all dashboard endpoints."""

    app = FastAPI(
        title="Quant Post-Trade Dashboard",
        description="Real-time post-trade analytics",
        version="0.2.0",
    )

    app.add_middleware(
        CORSMiddleware,
        allow_origins=["http://localhost:3000"],
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )

    @app.get("/health")
    def health():
        return {"status": "ok"}

    @app.get("/api/symbols")
    def active_symbols():
        """Return list of symbols with active positions or recent fills."""
        return state.get_active_symbols()

    # Tab 1: PnL Attribution
    @app.get("/api/pnl")
    def pnl_summary(symbol: str | None = Query(None)):
        return state.get_pnl_summary(symbol=symbol)

    # Tab 2: Transaction Cost Analysis
    @app.get("/api/tca")
    def tca_summary(symbol: str | None = Query(None)):
        return state.get_tca_summary(symbol=symbol)

    # Tab 3: Alpha Decay — IC at multiple time horizons
    @app.get("/api/alpha-decay")
    def alpha_decay(symbol: str | None = Query(None)):
        return state.get_alpha_decay(symbol=symbol)

    # Tab 4: Risk Metrics
    @app.get("/api/risk-metrics")
    def risk_metrics():
        return state.get_risk_metrics()

    # Tab 5: Drawdown Analysis
    @app.get("/api/drawdown")
    def drawdown():
        return state.get_drawdown_data()

    # Tab 6: Fill Analysis
    @app.get("/api/fills")
    def fill_analysis(symbol: str | None = Query(None)):
        return state.get_fill_analysis(symbol=symbol)

    # Excel Export
    @app.get("/api/export/excel")
    def export_excel():
        data = state.get_all_data_for_export()
        workbook_bytes = _build_excel(data)
        return StreamingResponse(
            io.BytesIO(workbook_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=post_trade_report.xlsx"},
        )

    # ------------------------------------------------------------------
    # Tab 7: Backtest Analysis — async job submission + polling
    # ------------------------------------------------------------------

    job_store = JobStore(max_workers=2)

    @app.post("/api/analysis/submit")
    def submit_analysis(req: AnalysisRequest):
        """Submit an analysis job. Returns job_id for polling."""
        job_id = job_store.submit(req.analysis_type, req.params)
        return {"job_id": job_id}

    @app.get("/api/analysis/status/{job_id}")
    def job_status(job_id: str):
        """Poll job status and progress."""
        job = job_store.get(job_id)
        if not job:
            return {"error": "Job not found"}
        return {
            "job_id": job.job_id,
            "analysis_type": job.analysis_type,
            "status": job.status.value,
            "progress": job.progress,
            "error": job.error,
            "has_result": job.result is not None,
        }

    @app.get("/api/analysis/result/{job_id}")
    def job_result(job_id: str):
        """Retrieve completed job result."""
        job = job_store.get(job_id)
        if not job:
            return {"error": "Job not found"}
        if job.result is None:
            return {"error": "Job not yet completed", "status": job.status.value}
        return {
            "job_id": job.job_id,
            "analysis_type": job.analysis_type,
            "status": job.status.value,
            "result": job.result,
        }

    @app.get("/api/analysis/backtests")
    def list_backtests():
        """List historical backtest runs that have trade data available."""
        return {"backtests": _list_historical_backtests()}

    @app.get("/api/analysis/db-symbols")
    def list_db_symbols():
        """List symbols in TimescaleDB with their trade counts and date range."""
        import os

        import psycopg2

        dsn = os.getenv("DATABASE_URL", "postgresql://quant:quant_dev@timescaledb:5432/quantdb")
        with psycopg2.connect(dsn) as conn, conn.cursor() as cur:
            cur.execute(
                "SELECT symbol, count(*), "
                "extract(epoch from min(time)) * 1000, "
                "extract(epoch from max(time)) * 1000 "
                "FROM trades "
                "WHERE backtest_id IS NULL "
                "GROUP BY symbol ORDER BY count(*) DESC"
            )
            rows = cur.fetchall()
        return {
            "symbols": [
                {
                    "symbol": r[0],
                    "count": int(r[1]),
                    "start_ms": int(r[2]),
                    "end_ms": int(r[3]),
                }
                for r in rows
            ]
        }

    @app.get("/api/analysis/jobs")
    def list_jobs(limit: int = Query(20)):
        """List recent analysis jobs."""
        jobs = job_store.list_jobs(limit=limit)
        return {
            "jobs": [
                {
                    "job_id": j.job_id,
                    "analysis_type": j.analysis_type,
                    "status": j.status.value,
                    "progress": j.progress,
                    "created_at": j.created_at,
                    "completed_at": j.completed_at,
                    "has_result": j.result is not None,
                    "error": j.error,
                }
                for j in jobs
            ]
        }

    return app


def _build_excel(data: dict) -> bytes:
    """Build an Excel workbook with one sheet per tab."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")

    def write_header(ws, headers: list[str], row: int = 1) -> None:
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill

    # --- Sheet 1: PnL Summary ---
    ws_pnl = wb.active
    ws_pnl.title = "PnL Summary"
    pnl = data["pnl"]
    summary_rows = [
        ("Initial Equity", pnl["initial_equity"]),
        ("Current Equity", pnl["current_equity"]),
        ("Total Realized PnL", pnl["total_realized_pnl"]),
        ("Total Unrealized PnL", pnl["total_unrealized_pnl"]),
        ("Total Fees", pnl["total_fees"]),
        ("Total Return %", pnl["total_return_pct"]),
        ("Number of Fills", pnl["num_fills"]),
    ]
    write_header(ws_pnl, ["Metric", "Value"])
    for i, (metric, value) in enumerate(summary_rows, 2):
        ws_pnl.cell(row=i, column=1, value=metric)
        ws_pnl.cell(row=i, column=2, value=value)

    # Positions
    if pnl["positions"]:
        row = len(summary_rows) + 4
        ws_pnl.cell(row=row - 1, column=1, value="Positions").font = header_font
        write_header(ws_pnl, ["Symbol", "Quantity", "Avg Entry", "Realized PnL", "Unrealized PnL", "Fees"], row)
        for sym, pos in pnl["positions"].items():
            row += 1
            ws_pnl.cell(row=row, column=1, value=sym)
            ws_pnl.cell(row=row, column=2, value=pos["quantity"])
            ws_pnl.cell(row=row, column=3, value=pos["avg_entry_price"])
            ws_pnl.cell(row=row, column=4, value=pos["realized_pnl"])
            ws_pnl.cell(row=row, column=5, value=pos["unrealized_pnl"])
            ws_pnl.cell(row=row, column=6, value=pos["total_fees"])

    ws_pnl.column_dimensions["A"].width = 22
    ws_pnl.column_dimensions["B"].width = 18

    # --- Sheet 2: TCA ---
    ws_tca = wb.create_sheet("TCA")
    tca = data["tca"]
    if tca.get("fills"):
        headers = [
            "Fill ID",
            "Symbol",
            "Side",
            "Spread (bps)",
            "Slippage (bps)",
            "Market Impact (bps)",
            "Fee (bps)",
            "Total Cost (bps)",
        ]
        write_header(ws_tca, headers)
        for i, f in enumerate(tca["fills"], 2):
            ws_tca.cell(row=i, column=1, value=f["fill_id"][:12])
            ws_tca.cell(row=i, column=2, value=f["symbol"])
            ws_tca.cell(row=i, column=3, value=f["side"])
            ws_tca.cell(row=i, column=4, value=f["spread_cost_bps"])
            ws_tca.cell(row=i, column=5, value=f["slippage_bps"])
            ws_tca.cell(row=i, column=6, value=f["market_impact_bps"])
            ws_tca.cell(row=i, column=7, value=f["fee_bps"])
            ws_tca.cell(row=i, column=8, value=f["total_cost_bps"])

        # Averages row
        row = len(tca["fills"]) + 3
        ws_tca.cell(row=row, column=1, value="AVERAGES").font = header_font
        avgs = tca["averages"]
        ws_tca.cell(row=row, column=4, value=avgs["spread_cost_bps"])
        ws_tca.cell(row=row, column=5, value=avgs["slippage_bps"])
        ws_tca.cell(row=row, column=6, value=avgs["market_impact_bps"])
        ws_tca.cell(row=row, column=7, value=avgs["fee_bps"])
        ws_tca.cell(row=row, column=8, value=avgs["total_cost_bps"])

    # --- Sheet 3: Alpha Decay ---
    ws_alpha = wb.create_sheet("Alpha Decay")
    alpha = data["alpha_decay"]
    write_header(ws_alpha, ["Horizon", "IC", "Filled Signals", "Total Signals"])
    for i, h in enumerate(alpha.get("horizons", []), 2):
        ws_alpha.cell(row=i, column=1, value=h["horizon_label"])
        ws_alpha.cell(row=i, column=2, value=h["ic"])
        ws_alpha.cell(row=i, column=3, value=h["filled_count"])
        ws_alpha.cell(row=i, column=4, value=h["total_signals"])
    # Per-strategy breakdown
    row = len(alpha.get("horizons", [])) + 4
    for strat_id, strat_data in alpha.get("strategies", {}).items():
        ws_alpha.cell(row=row, column=1, value=strat_id).font = header_font
        ws_alpha.cell(row=row, column=2, value=f"({strat_data['signal_count']} signals)")
        row += 1
        write_header(ws_alpha, ["Horizon", "IC", "Filled Signals"], row)
        row += 1
        for h in strat_data["horizons"]:
            ws_alpha.cell(row=row, column=1, value=h["horizon_label"])
            ws_alpha.cell(row=row, column=2, value=h["ic"])
            ws_alpha.cell(row=row, column=3, value=h["filled_count"])
            row += 1
        row += 1
    ws_alpha.column_dimensions["A"].width = 18
    ws_alpha.column_dimensions["B"].width = 14

    # --- Sheet 4: Risk Metrics ---
    ws_risk = wb.create_sheet("Risk Metrics")
    metrics = data["risk_metrics"]
    write_header(ws_risk, ["Metric", "Value"])
    metric_rows = [
        ("Sharpe Ratio", metrics["sharpe_ratio"]),
        ("Sortino Ratio", metrics["sortino_ratio"]),
        ("Calmar Ratio", metrics["calmar_ratio"]),
        ("Max Drawdown %", metrics["max_drawdown_pct"]),
        ("Max DD Duration (periods)", metrics["max_drawdown_duration_periods"]),
        ("Total Return %", metrics["total_return_pct"]),
        ("Annualized Return %", metrics["annualized_return_pct"]),
        ("Win Rate %", metrics["win_rate_pct"]),
        ("Profit Factor", metrics["profit_factor"]),
        ("Number of Trades", metrics["num_trades"]),
        ("Current Equity", metrics["current_equity"]),
        ("Peak Equity", metrics["peak_equity"]),
    ]
    for i, (metric, value) in enumerate(metric_rows, 2):
        ws_risk.cell(row=i, column=1, value=metric)
        ws_risk.cell(row=i, column=2, value=value)
    ws_risk.column_dimensions["A"].width = 28
    ws_risk.column_dimensions["B"].width = 18

    # --- Sheet 4: Drawdown ---
    ws_dd = wb.create_sheet("Drawdown")
    dd = data["drawdown"]
    write_header(ws_dd, ["Timestamp", "Equity", "Drawdown %"])
    for i, (eq, ddc) in enumerate(zip(dd["equity_curve"], dd["drawdown_curve"], strict=False), 2):
        ws_dd.cell(row=i, column=1, value=eq["timestamp"])
        ws_dd.cell(row=i, column=2, value=eq["equity"])
        ws_dd.cell(row=i, column=3, value=ddc["drawdown_pct"])

    # --- Sheet 5: Fills ---
    ws_fills = wb.create_sheet("Fills")
    fills = data["fills"]
    if fills.get("fills"):
        headers = [
            "Fill ID",
            "Timestamp",
            "Symbol",
            "Side",
            "Quantity",
            "Price",
            "Fee",
            "Slippage (bps)",
            "Strategy",
            "Mode",
        ]
        write_header(ws_fills, headers)
        for i, f in enumerate(fills["fills"], 2):
            ws_fills.cell(row=i, column=1, value=f["fill_id"][:12])
            ws_fills.cell(row=i, column=2, value=f["timestamp"])
            ws_fills.cell(row=i, column=3, value=f["symbol"])
            ws_fills.cell(row=i, column=4, value=f["side"])
            ws_fills.cell(row=i, column=5, value=f["quantity"])
            ws_fills.cell(row=i, column=6, value=f["fill_price"])
            ws_fills.cell(row=i, column=7, value=f["fee"])
            ws_fills.cell(row=i, column=8, value=f["slippage_bps"])
            ws_fills.cell(row=i, column=9, value=f["strategy_id"])
            ws_fills.cell(row=i, column=10, value=f.get("trading_mode", "paper"))

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
